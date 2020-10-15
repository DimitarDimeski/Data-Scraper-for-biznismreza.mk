#! python3
# Data Scraper for the website biznismreza.mk
from selenium import webdriver 
from selenium.webdriver.common.keys import Keys
import time
import openpyxl


# Open excel and sheet

wb = openpyxl.load_workbook('example_excel.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')



n= int(input('From which cell do I start?')) #because the first row is header 

cells = {}

def one_row_down(cells): 
    global n 
    n = n+1
    cells['embs']= sheet.cell(row = n,column = 3)
    cells['revenue2019']= sheet.cell(row=n, column = 12)
    cells['year_of_estbl']=sheet.cell(row=n, column = 4)
    cells['sector'] = sheet.cell(row=n, column = 10)
    cells['revenue2018']= sheet.cell(row=n, column = 11)
    cells['jobs2018'] = sheet.cell(row=n, column = 13)
    cells['jobs2019'] = sheet.cell(row=n, column = 14)
    cells['names'] = sheet.cell(row=n, column = 16) 
    
one_row_down(cells) #create the dictionary

browser = webdriver.Chrome('D:\\Downloads\\chromedriver') # go to website
browser.get('https://biznismreza.mk/Account/Login') 

browser.maximize_window()

#Login (input user and pass):
username = input('Username: ')
password =  input('Password: ')
username_field = browser.find_element_by_id('Username') 
username_field.send_keys(username)
password_field = browser.find_element_by_id('Password') 
password_field.send_keys(password)
password_field.send_keys(Keys.ENTER)


# Open portfolio:
browser.get('https://biznismreza.mk/Management/Portfolio')
i=1 #starting point for iterations
iterations = input("What's the limit? ") #asking how many iterations

missing_embs=0
while i <= int(iterations): 
    
    print(f'Currently at: {i}') 
    
    # Check if filled out in excel (check revenues for 2019 cell)
    if(cells['revenue2019'].value != None or cells['embs'].value == None ):
        if cells['embs'].value == None:
            missing_embs = missing_embs + 1
             
        
        if missing_embs == 10:
            i=50
            break
        
        print('Passing') #pass if 2019 revenues are already filled
        print(cells['embs'].value)
        one_row_down(cells)
        continue
    
    missing_embs = 0
    
    print('Found')
    # Search company:
    company = browser.find_element_by_id("MainContent_EMBS")
    company.send_keys(str(cells['embs'].value)) #copy the embs from the excel to pass to the webdriver
    browser.find_element_by_id("MainContent_SearchEmbs").click()
    browser.implicitly_wait(5)
    try: 
        browser.find_element_by_link_text(str(cells['embs'].value)).click() 
    except Exception: #if embs not correct, skip company and print embs_error
        print('embs_error')
        cells['names'].value = 'embs_error'
        one_row_down(cells)
        browser.refresh()
        continue
    
    
    # switch tabs:
    
    browser.switch_to.window(browser.window_handles[1])
    
    # Copy basic data (sector and year of establishment):
    browser.implicitly_wait(5)
    dejnost = browser.find_element_by_id("MainContent_GlavnaPrihodnaShifra")
    print(f"dejnost:{dejnost.text}")
    cells['sector'].value = dejnost.text
    
    time.sleep(4)
    
    osnovana = browser.find_element_by_id('MainContent_DatumNaOsnovanje')
    browser.execute_script("arguments[0].scrollIntoView();", osnovana)
    print(f'osnovana:{osnovana.text}')
    cells['year_of_estbl'].value = osnovana.text
    
    
    sopsUprv=[]
    try:
        upraviteli = browser.find_elements_by_id('divUpraviteli')
        
        for upravitel in upraviteli:
            sopsUprv.append(upravitel.text)
            
        sopstvenici = browser.find_elements_by_id('divSopstvenici')
        
        for sopstevnik in sopstvenici:
            sopsUprv.append(sopstevnik.text)
    except Exception:
        pass
        
    print(f"upraviteli:{sopsUprv}")
    cells['names'].value = repr(sopsUprv)
    
    
    # Check if data point for 2019 available:
    
    if (browser.find_element_by_id("MainContent_FinansiskiPodatociUc1_lblCurrentYear").text == '2019'):
        prihodi2019 = browser.find_element_by_id('MainContent_FinansiskiPodatociUc1_lblVkupniPrihodiCurrentYear').text
        print(f"prihodi_za_2019:{prihodi2019}")
        cells['revenue2019'].value = prihodi2019.replace('.','')
        
        prihodi2018 = browser.find_element_by_id('MainContent_FinansiskiPodatociUc1_lblVkupniPrihodiLastYear').text
        print(f"prihodi_za_2018:{prihodi2018}")
        cells['revenue2018'].value = prihodi2018.replace('.','')
        
        browser.find_element_by_id("ctl06_AdditionalMenuPlaceHolder_AdditionalMenu1_FinansiiLink").click()
        browser.implicitly_wait(5)
        
        vraboteni2019 = browser.find_element_by_id('MainContent_lblProsecenBrojNaVraboteniYear1').text
        print(f"vraboteni_za_2019:{vraboteni2019}")
        cells['jobs2019'].value = vraboteni2019
        
        vraboteni2018 = browser.find_element_by_id('MainContent_lblProsecenBrojNaVraboteniYear2').text
        print(f"vraboteni_za_2018:{vraboteni2018}")
        cells['jobs2018'].value = vraboteni2018
    else: 
        browser.find_element_by_id("ctl06_AdditionalMenuPlaceHolder_AdditionalMenu1_FinansiiLink").click()
        browser.implicitly_wait(5)
    
    # Closing the current window and going back to portfolio:
        
    browser.close()
    browser.switch_to.window(browser.window_handles[0])
    browser.refresh()
    one_row_down(cells)
    wb.save('RJ-final.xlsx')
    print('\n')
    
    i+=1

wb.save('RJ-final.xlsx')
browser.close()

print ('All done!')

